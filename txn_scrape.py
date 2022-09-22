import time
import openpyxl
from selenium import webdriver
from requests_html import HTMLSession
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
wb = load_workbook(filename = 'txn_sheet.xlsx')
ws = wb.active
id_list = []
for row in range(1, 4): # Change the second value to the number of transactions + 1
    id_list.append(ws['A' + str(row)].value)
# For Roland: You are going to have to download chromedriver from the following website: https://sites.google.com/chromium.org/driver/downloads. Ensure the version you choose corresponds to the version of chrome you are currently running. Extract the file to the PATH variable shown below.
PATH = 'C:\Program Files (x86)\chromedriver.exe'
signer_list = []
counter = 1
for id in id_list:
    url = 'https://explorer.solana.com/tx/' + id + '?cluster=testnet'
    browser = webdriver.Chrome(PATH)
    browser.get(url)  # Opens the web browser
    time.sleep(5)  # Waits for the transactions to load, you may have to increase this value if Transactions fail to load in time.
    res = browser.find_elements(By.CLASS_NAME, "font-monospace")
    text = res[4].text
    print(str(counter)+'.'+text)
    counter+=1
    signer_list.append(text)
print(signer_list)
count = 1
for signer_id in signer_list:
    ws['B' + str(count)] = signer_id
    count+=1
wb.save('txn_sheet.xlsx')